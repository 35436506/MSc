import random
import math
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt

# Define the recipes as a list from 1 to 65
recipes = list(range(1, 66))

# Define the recipe eligibility for each factory
factory_eligibility = {
    'F1': random.sample(recipes, int(0.7 * len(recipes))),  # 70% of total recipes are eligible for F1
    'F2': random.sample(recipes, int(0.9 * len(recipes))),  # 90% of total recipes are eligible for F2
    'F3': recipes  # 100% recipes are eligible for F3
}

# Generate random orders with 1-4 recipes
def generate_orders(num_orders, is_real=True):
    orders = []
    for n in range(num_orders):
        recipe_ids = random.sample(recipes, random.randint(1, 4))
        orders.append({'id': len(orders) + 1, 'recipe_ids': recipe_ids, 'is_real': is_real})
    return orders

# Get the eligible factories for each order based on the recipes
def get_eligible_factories(order):
    eligible_factories = []
    for factory, eligible_recipes in factory_eligibility.items():
        if all(recipe_id in eligible_recipes for recipe_id in order['recipe_ids']):
            eligible_factories.append(factory)
    return eligible_factories

# Generate sample orders for day t-1 and day t
num_total_orders = 600
num_real_orders_t_minus_1 = num_total_orders // 2
num_simulated_orders_t_minus_1 = num_total_orders - num_real_orders_t_minus_1
num_new_real_orders_t = num_total_orders // 6
num_simulated_orders_t = num_total_orders // 3

orders_t_minus_1 = generate_orders(num_real_orders_t_minus_1) + generate_orders(num_simulated_orders_t_minus_1, is_real=False)
orders_t = orders_t_minus_1[:num_real_orders_t_minus_1] + generate_orders(num_new_real_orders_t) + generate_orders(num_simulated_orders_t, is_real=False)

# Re-number the orders for day t-1 and day t
for i, order in enumerate(orders_t_minus_1):
    order['id'] = i + 1

for i, order in enumerate(orders_t):
    order['id'] = i + 1

# Add eligible factories to each order
for order in orders_t_minus_1 + orders_t:
    order['eligible_factories'] = get_eligible_factories(order)

# Define factory capacities
factory_capacities = {
    'F1': int((num_total_orders) / 3),  # Serve a third of total orders
    'F2': int((num_total_orders) / 2),  # Serve half of total orders
    'F3': float('inf')  # Catch-all factory with infinite capacity
}

# Print the maximum capacity of each factory
print("Maximum capacity of each factory:")
for factory, capacity in factory_capacities.items():
    print(f"{factory}: {capacity}")

# Initial allocation (orders are allocated based on priority: F1 -> F2 -> F3)
def initial_allocation(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    # Allocate orders to F1
    for order in remaining_orders.copy():
        if 'F1' in order['eligible_factories'] and len(allocation['F1']) < factory_capacities['F1']:
            allocation['F1'].append(order)
            remaining_orders.remove(order)

    # Allocate orders to F2
    for order in remaining_orders.copy():
        if 'F2' in order['eligible_factories'] and len(allocation['F2']) < factory_capacities['F2']:
            allocation['F2'].append(order)
            remaining_orders.remove(order)

    # Allocate remaining orders to F3
    allocation['F3'].extend(remaining_orders)

    return allocation

# Perform initial allocation for day t-1 and day t
allocation_t_minus_1 = initial_allocation(orders_t_minus_1, factory_capacities)
allocation_t = initial_allocation(orders_t, factory_capacities)

# Plot solution allocation for day t-1
plt.figure(figsize=(8, 4))
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]
simulated_orders_t_minus_1 = [order for order in orders_t_minus_1 if not order['is_real']]
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t_minus_1['F1'] if order['is_real']]),
                             len([order for order in allocation_t_minus_1['F2'] if order['is_real']]),
                             len([order for order in allocation_t_minus_1['F3'] if order['is_real']])],
        color='blue', label='Real orders')
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t_minus_1['F1'] if not order['is_real']]),
                             len([order for order in allocation_t_minus_1['F2'] if not order['is_real']]),
                             len([order for order in allocation_t_minus_1['F3'] if not order['is_real']])],
        color='yellow', label='Simulated orders', bottom=[len([order for order in allocation_t_minus_1[factory] if order['is_real']]) for factory in ['F1', 'F2', 'F3']])
plt.xlabel('Factory')
plt.ylabel('Volume')
plt.title('Allocation for Day t-1')
plt.legend()
plt.show()

# Plot solution allocation for day t before SA
plt.figure(figsize=(8, 4))
real_orders_t = [order for order in orders_t if order['is_real']]
simulated_orders_t = [order for order in orders_t if not order['is_real']]
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t['F1'] if order['is_real']]),
                             len([order for order in allocation_t['F2'] if order['is_real']]),
                             len([order for order in allocation_t['F3'] if order['is_real']])],
        color='blue', label='Real orders')
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t['F1'] if not order['is_real']]),
                             len([order for order in allocation_t['F2'] if not order['is_real']]),
                             len([order for order in allocation_t['F3'] if not order['is_real']])],
        color='yellow', label='Simulated orders', bottom=[len([order for order in allocation_t[factory] if order['is_real']]) for factory in ['F1', 'F2', 'F3']])
plt.xlabel('Factory')
plt.ylabel('Volume')
plt.title('Allocation for Day t (Before SA)')
plt.legend()
plt.show()

# Calculate WMAPE site
def calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_t_items = 0
    for factory in ['F1', 'F2']:
        recipe_counts_t_minus_1 = {}
        recipe_counts_t = {}
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
        sheet_wmape.append([f"Recipe counts for {factory} on day t-1: {recipe_counts_t_minus_1}"])
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
        sheet_wmape.append([f"Recipe counts for {factory} on day t: {recipe_counts_t}"])
        for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
            t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
            t_count = recipe_counts_t.get(recipe_id, 0)
            abs_diff = abs(t_minus_1_count - t_count)
            total_abs_diff += abs_diff
            sheet_wmape.append([f"Recipe {recipe_id}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
    if total_t_items == 0:
        wmape_site = float('inf')
    else:
        wmape_site = total_abs_diff / total_t_items
    sheet_wmape.append([f"Total absolute difference: {total_abs_diff}"])
    sheet_wmape.append([f"Total items on day t: {total_t_items}"])
    sheet_wmape.append([f"WMAPE site: {wmape_site:.2f}"])
    return wmape_site

# Calculate WMAPE global
def calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    sheet_wmape.append([f"Recipe counts on day t-1: {recipe_counts_t_minus_1}"])
    for factory in allocation_t:
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    sheet_wmape.append([f"Recipe counts on day t: {recipe_counts_t}"])
    for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
        sheet_wmape.append([f"Recipe {recipe_id}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
    wmape_global = total_abs_diff / total_t_items
    sheet_wmape.append([f"Total absolute difference: {total_abs_diff}"])
    sheet_wmape.append([f"Total items on day t: {total_t_items}"])
    sheet_wmape.append([f"WMAPE global: {wmape_global:.2f}"])
    return wmape_global

# Simulated Annealing
def simulated_annealing(allocation_t_minus_1, allocation_t, factory_capacities, initial_temp=1000, cooling_rate=0.95, iterations=1000):
    current_allocation = {factory: orders[:] for factory, orders in allocation_t.items()}
    current_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, sheet_wmape)
    best_wmape_site = current_wmape_site
    temp = initial_temp

    for i in range(iterations):
        factory_from, factory_to = random.sample(list(factory_capacities.keys()), 2)
        if len(current_allocation[factory_from]) > 0 and len(current_allocation[factory_to]) < factory_capacities[factory_to]:
            order_from_index = random.randint(0, len(current_allocation[factory_from]) - 1)
            order_from = current_allocation[factory_from][order_from_index]
            if factory_to in order_from['eligible_factories']:
                current_allocation[factory_from].pop(order_from_index)
                current_allocation[factory_to].append(order_from)
                new_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, sheet_wmape)
                delta = new_wmape_site - current_wmape_site
                if delta < 0 or random.random() < math.exp(-delta / temp):
                    current_wmape_site = new_wmape_site
                    if current_wmape_site < best_wmape_site:
                        best_wmape_site = current_wmape_site
                else:
                    current_allocation[factory_to].pop()
                    current_allocation[factory_from].insert(order_from_index, order_from)
        temp *= cooling_rate

    # Ensure capacities of F1 and F2 are fully met while minimizing WMAPE site
    for factory in ['F1', 'F2']:
        while len(current_allocation[factory]) < factory_capacities[factory]:
            if current_allocation['F3']:
                best_order = None
                best_wmape_site_change = float('inf')
                for i, order in enumerate(current_allocation['F3']):
                    if all(recipe_id in factory_eligibility[factory] for recipe_id in order['recipe_ids']):
                        current_allocation[factory].append(order)
                        new_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, sheet_wmape)
                        wmape_site_change = new_wmape_site - current_wmape_site
                        if wmape_site_change < best_wmape_site_change:
                            best_order = i
                            best_wmape_site_change = wmape_site_change
                        current_allocation[factory].pop()
                if best_order is not None:
                    current_allocation[factory].append(current_allocation['F3'][best_order])
                    del current_allocation['F3'][best_order]
                    current_wmape_site += best_wmape_site_change
            else:
                break

    return current_allocation, current_wmape_site

# Export data and allocation decisions to Excel
workbook = Workbook()

# Create a new sheet for WMAPE calculations
sheet_wmape = workbook.create_sheet("WMAPE")

start_time = time.time()  # Start measuring execution time

# Calculate WMAPE site and global before SA
wmape_site_before_sa = calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape)
wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape)

print(f"WMAPE site before SA: {wmape_site_before_sa:.2f}")
print(f"WMAPE global: {wmape_global:.2f}")

# Export initial allocation for day t
sheet_t_allocation = workbook.create_sheet("Day t Allocation")
sheet_t_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation.append([factory, ", ".join(map(str, order_ids))])

# Apply Simulated Annealing to optimize allocation on day t
optimized_allocation_t, optimized_wmape_site = simulated_annealing(allocation_t_minus_1, allocation_t, factory_capacities)

# Plot solution allocation for day t after SA
plt.figure(figsize=(8, 4))
real_orders_t_after_sa = [order for order in optimized_allocation_t['F1'] + optimized_allocation_t['F2'] + optimized_allocation_t['F3'] if order['is_real']]
simulated_orders_t_after_sa = [order for order in optimized_allocation_t['F1'] + optimized_allocation_t['F2'] + optimized_allocation_t['F3'] if not order['is_real']]
plt.bar(['F1', 'F2', 'F3'], [len([order for order in optimized_allocation_t['F1'] if order['is_real']]),
                             len([order for order in optimized_allocation_t['F2'] if order['is_real']]),
                             len([order for order in optimized_allocation_t['F3'] if order['is_real']])],
        color='blue', label='Real orders')
plt.bar(['F1', 'F2', 'F3'], [len([order for order in optimized_allocation_t['F1'] if not order['is_real']]),
                             len([order for order in optimized_allocation_t['F2'] if not order['is_real']]),
                             len([order for order in optimized_allocation_t['F3'] if not order['is_real']])],
        color='yellow', label='Simulated orders', bottom=[len([order for order in optimized_allocation_t[factory] if order['is_real']]) for factory in ['F1', 'F2', 'F3']])
plt.xlabel('Factory')
plt.ylabel('Volume')
plt.title('Allocation for Day t (After SA)')
plt.legend()
plt.show()

# Calculate WMAPE site after SA
wmape_site_after_sa = calculate_wmape_site(allocation_t_minus_1, optimized_allocation_t, sheet_wmape)
print(f"WMAPE site after SA: {wmape_site_after_sa:.2f}")

end_time = time.time()  # End measuring execution time
execution_time = end_time - start_time

print(f"Execution time: {execution_time:.2f} seconds")

# Export orders for day t-1
sheet_t_minus_1_orders = workbook.create_sheet("Day t-1 Orders")
sheet_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t_minus_1:
    sheet_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Export allocation for day t-1
sheet_t_minus_1_allocation = workbook.create_sheet("Day t-1 Allocation")
sheet_t_minus_1_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t_minus_1.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_minus_1_allocation.append([factory, ", ".join(map(str, order_ids))])

# Export orders for day t
sheet_t_orders = workbook.create_sheet("Day t Orders")
sheet_t_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t:
    sheet_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Export allocation for day t after SA
sheet_t_allocation_sa = workbook.create_sheet("Day t Allocation (SA)")
sheet_t_allocation_sa.append(["Factory", "Allocated Orders"])
for factory, orders in optimized_allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation_sa.append([factory, ", ".join(map(str, order_ids))])

# Remove the default blank sheet
default_sheet = workbook['Sheet']
workbook.remove(default_sheet)

workbook.save("allocation_results.xlsx")