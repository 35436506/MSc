import random
import math
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

# Set seed for reproducibility
random.seed(42)

# Define the recipes as a list from 1 to 65
recipes = list(range(1, 66))

# Define the recipe eligibility for each factory
factory_eligibility = {
    'F1': random.sample(recipes, int(0.3 * len(recipes))),  # 30% of recipes
    'F2': random.sample(recipes, int(0.6 * len(recipes))),  # 60% of recipes
    'F3': recipes  # All recipes
}

# Define total orders and factory capacities
total_orders = 1000
F1_cap = int(0.3 * total_orders)  # 30% of total orders
F2_cap = int(0.5 * total_orders)  # 50% of total orders
factory_capacities = {
    'F1': F1_cap,
    'F2': F2_cap,
    'F3': float('inf')  # F3 has unlimited capacity
}

def generate_orders(num_orders, is_real=True, factory=None, prioritize_f1_f2=False):
    orders = []
    while len(orders) < num_orders:
        if prioritize_f1_f2:
            if factory == 'F1':
                recipe_ids = random.sample(factory_eligibility['F1'], random.randint(1, 4))
            elif factory == 'F2':
                recipe_ids = random.sample(factory_eligibility['F2'], random.randint(1, 4))
            else:
                recipe_ids = random.sample(recipes, random.randint(1, 4))
        else:
            recipe_ids = random.sample(recipes, random.randint(1, 4))

        if factory is None or all(recipe_id in factory_eligibility[factory] for recipe_id in recipe_ids):
            order = {'id': len(orders) + 1, 'recipe_ids': recipe_ids, 'is_real': is_real}
            order['eligible_factories'] = get_eligible_factories(order)
            orders.append(order)
    return orders

def get_eligible_factories(order):
    return [factory for factory, eligible_recipes in factory_eligibility.items()
            if all(recipe_id in eligible_recipes for recipe_id in order['recipe_ids'])]

def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    num_real = int(real_proportion * total_orders)
    num_simulated = int(simulated_proportion * total_orders)
    
    if existing_real_orders:
        orders = existing_real_orders.copy()
        additional_real_orders = num_real - len(existing_real_orders)
        if additional_real_orders > 0:
            orders.extend(generate_orders(additional_real_orders))
    else:
        orders = generate_orders(num_real)
    
    orders.extend(generate_orders(num_simulated, is_real=False))
    
    for factory in ['F1', 'F2']:
        eligible_orders = [order for order in orders if factory in order['eligible_factories']]
        while len(eligible_orders) < factory_capacities[factory]:
            new_orders = generate_orders(factory_capacities[factory] - len(eligible_orders), 
                                         is_real=False, factory=factory, prioritize_f1_f2=True)
            orders.extend(new_orders)
            eligible_orders.extend(new_orders)
    
    for i, order in enumerate(orders):
        order['id'] = i + 1
    
    return orders

def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    for factory in ['F1', 'F2', 'F3']:
        capacity = factory_capacities[factory]
        eligible_orders = [order for order in remaining_orders if factory in order['eligible_factories']]
        eligible_orders.sort(key=lambda x: x['is_real'], reverse=True)  # Prioritize real orders
        
        if factory == 'F3':
            allocated_orders = eligible_orders  # Allocate all remaining orders to F3
        else:
            allocated_orders = eligible_orders[:int(capacity)]  # Convert capacity to int for slicing
        
        allocation[factory].extend(allocated_orders)
        remaining_orders = [order for order in remaining_orders if order not in allocated_orders]

        if factory == 'F2' and len(allocation[factory]) < capacity:
            # If F2 is not full, generate more orders
            additional_orders = generate_orders(int(capacity) - len(allocation[factory]), 
                                                is_real=False, factory='F2', prioritize_f1_f2=True)
            allocation[factory].extend(additional_orders)

    return allocation

def plot_allocation(allocation, day):
    plt.figure(figsize=(12, 6))
    factories = ['F1', 'F2', 'F3']
    real_order_counts = [len([order for order in allocation[factory] if order['is_real']]) for factory in factories]
    simulated_order_counts = [len([order for order in allocation[factory] if not order['is_real']]) for factory in factories]

    x = np.arange(len(factories))
    width = 0.35

    plt.bar(x, real_order_counts, width, label='Real orders', color='#1f77b4')
    plt.bar(x, simulated_order_counts, width, bottom=real_order_counts, label='Simulated orders', color='#ff7f0e')

    max_capacity_line = None
    for i, factory in enumerate(factories):
        capacity = factory_capacities[factory]
        if factory == 'F3':
            capacity = real_order_counts[i] + simulated_order_counts[i] + 1000
        
        line_width = 0.6
        xmin = i - line_width/2
        xmax = i + line_width/2
        
        line = plt.hlines(y=capacity, xmin=xmin, xmax=xmax, 
                    colors='r', linestyles='-', linewidth=2)
        if max_capacity_line is None:
            max_capacity_line = line

    plt.xlabel('Factory')
    plt.ylabel('Order quantity')
    plt.title(f'Allocation for Day {day}')
    plt.xticks(x, factories)
    
    plt.legend([plt.Rectangle((0,0),1,1,fc="#1f77b4"), 
                plt.Rectangle((0,0),1,1,fc="#ff7f0e"),
                max_capacity_line],
               ['Real orders', 'Simulated orders', 'Max capacity'],
               bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.show()

def calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_items_t = 0
    
    for factory in ['F1', 'F2', 'F3']:
        factory_abs_diff = 0
        recipe_counts_t_minus_1 = {}
        recipe_counts_t = {}
        
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
        
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_items_t += 1
        
        sheet_wmape.append([f"Recipe counts for {factory} on day t-1: {recipe_counts_t_minus_1}"])
        sheet_wmape.append([f"Recipe counts for {factory} on day t: {recipe_counts_t}"])
        
        for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
            t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
            t_count = recipe_counts_t.get(recipe_id, 0)
            abs_diff = abs(t_count - t_minus_1_count)
            factory_abs_diff += abs_diff
            sheet_wmape.append([f"Recipe {recipe_id} in {factory}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
        
        total_abs_diff += factory_abs_diff
        
        sheet_wmape.append([f"Total absolute difference for {factory}: {factory_abs_diff}"])
        sheet_wmape.append([f"Total items for {factory} on day t: {sum(recipe_counts_t.values())}"])
        
    if total_items_t == 0:
        wmape_site = float('inf')
    else:
        wmape_site = total_abs_diff / total_items_t
    
    sheet_wmape.append([f"Total absolute difference across all factories: {total_abs_diff}"])
    sheet_wmape.append([f"Total items across all factories on day t: {total_items_t}"])
    sheet_wmape.append([f"WMAPE site: {wmape_site:.3f}"])
    
    return wmape_site

def calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    sheet_wmape.append([f"Recipe counts on day t-1: {recipe_counts_t_minus_1}"])
    for factory in allocation_t:
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    sheet_wmape.append([f"Recipe counts on day t: {recipe_counts_t}"])
    for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
        sheet_wmape.append([f"Recipe {recipe_id}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
    wmape_global = total_abs_diff / total_t_items
    sheet_wmape.append([f"Total absolute difference: {total_abs_diff}"])
    sheet_wmape.append([f"Total items on day t: {total_t_items}"])
    sheet_wmape.append([f"WMAPE global: {wmape_global:.4f}"])
    return wmape_global

def simulated_annealing_with_swap(allocation_t_minus_1, allocation_t, factory_capacities, initial_temp=1000, cooling_rate=0.995, iterations=10000):
    current_allocation = {factory: orders[:] for factory, orders in allocation_t.items()}
    current_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, [])
    best_allocation = current_allocation.copy()
    best_wmape_site = current_wmape_site
    temp = initial_temp

    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1)

    for i in range(iterations):
        factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
        if current_allocation[factory1] and current_allocation[factory2]:
            idx1 = random.randint(0, len(current_allocation[factory1]) - 1)
            idx2 = random.randint(0, len(current_allocation[factory2]) - 1)
            order1 = current_allocation[factory1][idx1]
            order2 = current_allocation[factory2][idx2]
            
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                new_allocation = {f: orders[:] for f, orders in current_allocation.items()}
                new_allocation[factory1][idx1], new_allocation[factory2][idx2] = order2, order1
                
                new_wmape_site = calculate_wmape_site(allocation_t_minus_1, new_allocation, [])
                
                if new_wmape_site < current_wmape_site or random.random() < math.exp((current_wmape_site - new_wmape_site) / temp):
                    current_allocation = new_allocation
                    current_wmape_site = new_wmape_site
                    if current_wmape_site < best_wmape_site:
                        best_wmape_site = current_wmape_site
                        best_allocation = {f: orders[:] for f, orders in current_allocation.items()}
        
        # Local search
        if i % 100 == 0:
            current_allocation = local_search(allocation_t_minus_1, current_allocation, factory_capacities, recipe_counts_t_minus_1)
            current_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, [])
            if current_wmape_site < best_wmape_site:
                best_wmape_site = current_wmape_site
                best_allocation = current_allocation.copy()

        temp *= cooling_rate

    return best_allocation, best_wmape_site

def get_recipe_counts(allocation):
    recipe_counts = {}
    for factory in allocation:
        for order in allocation[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts[recipe_id] = recipe_counts.get(recipe_id, 0) + 1
    return recipe_counts

def local_search(allocation_t_minus_1, current_allocation, factory_capacities, recipe_counts_t_minus_1):
    recipe_counts_current = get_recipe_counts(current_allocation)
    
    for l in range(100):  # Perform 100 local search iterations
        factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
        if current_allocation[factory1] and current_allocation[factory2]:
            order1 = max(current_allocation[factory1], key=lambda o: sum(abs(recipe_counts_current.get(r, 0) - recipe_counts_t_minus_1.get(r, 0)) for r in o['recipe_ids']))
            order2 = max(current_allocation[factory2], key=lambda o: sum(abs(recipe_counts_current.get(r, 0) - recipe_counts_t_minus_1.get(r, 0)) for r in o['recipe_ids']))
            
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                idx1, idx2 = current_allocation[factory1].index(order1), current_allocation[factory2].index(order2)
                current_allocation[factory1][idx1], current_allocation[factory2][idx2] = order2, order1
                recipe_counts_current = get_recipe_counts(current_allocation)
    
    return current_allocation

# Main execution code
workbook = Workbook()

# Create sheets in the desired order
sheet_wmape = workbook.create_sheet("WMAPE", 0)
sheet_t_minus_1_orders = workbook.create_sheet("Day t-1 Orders", 1)
sheet_t_minus_1_allocation = workbook.create_sheet("Day t-1 Allocation", 2)
sheet_t_orders = workbook.create_sheet("Day t Orders", 3)
sheet_t_allocation = workbook.create_sheet("Day t Allocation (Initial)", 4)
sheet_t_allocation_sa = workbook.create_sheet("Day t Allocation (SA)", 5)

start_time = time.time()

# Generate orders for day t-1
orders_t_minus_1 = generate_orders_for_day(0.3, 0.7)

# Extract real orders from day t-1
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

# Calculate the number of new real orders for day t (10% of total orders)
new_real_orders_count = int(0.1 * total_orders)

# Calculate the total number of real orders for day t
total_real_orders_t = len(real_orders_t_minus_1) + new_real_orders_count

# Calculate the proportion of real orders for day t
real_proportion_t = total_real_orders_t / total_orders

# Generate orders for day t
orders_t = generate_orders_for_day(real_proportion_t, 1 - real_proportion_t, existing_real_orders=real_orders_t_minus_1)

# Perform allocation for day t-1 and day t
allocation_t_minus_1 = allocate_orders(orders_t_minus_1, factory_capacities)
allocation_t = allocate_orders(orders_t, factory_capacities)

# Plot allocations
plot_allocation(allocation_t_minus_1, 't-1')
plot_allocation(allocation_t, 't (Before SA)')

# Calculate WMAPE site and global before SA
wmape_site_before_sa = calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape)
wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape)

print(f"WMAPE site before SA: {wmape_site_before_sa:.3f}")
print(f"WMAPE global: {wmape_global:.3f}")

# Apply Simulated Annealing with swapping to optimize allocation on day t
optimized_allocation_t, optimized_wmape_site = simulated_annealing_with_swap(allocation_t_minus_1, allocation_t, factory_capacities)

# Export WMAPE site after SA to Excel sheet
sheet_wmape.append([f"WMAPE site after SA: {optimized_wmape_site:.3f}"])

# Plot solution allocation for day t after SA
plot_allocation(optimized_allocation_t, 't (After SA)')

print(f"WMAPE site after SA: {optimized_wmape_site:.3f}")

end_time = time.time()
execution_time = end_time - start_time

print(f"Execution time: {execution_time:.2f} seconds")
print("-------------------------------")
print("                               ")

# Populate Day t-1 Orders sheet
sheet_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t_minus_1:
    sheet_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t-1 Allocation sheet
sheet_t_minus_1_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t_minus_1.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_minus_1_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Orders sheet
sheet_t_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t:
    sheet_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t Allocation (Initial) sheet
sheet_t_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Allocation (SA) sheet
sheet_t_allocation_sa.append(["Factory", "Allocated Orders"])
for factory, orders in optimized_allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation_sa.append([factory, ", ".join(map(str, order_ids))])

# Remove the default sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Save the workbook
workbook.save("allocation_results.xlsx")



# Test scalability
def run_allocation_process(quantity):
    global total_orders, factory_capacities
    start_time = time.time()
    
    # Update total orders and factory capacities
    total_orders = quantity
    F1_cap = int(0.3 * total_orders)
    F2_cap = int(0.5 * total_orders)
    factory_capacities = {
        'F1': F1_cap,
        'F2': F2_cap,
        'F3': float('inf')
    }
    
    # Generate orders for day t-1
    orders_t_minus_1 = generate_orders_for_day(0.3, 0.7)

    # Extract real orders from day t-1
    real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

    # Calculate the number of new real orders for day t (10% of total orders)
    new_real_orders_count = int(0.1 * total_orders)

    # Calculate the total number of real orders for day t
    total_real_orders_t = len(real_orders_t_minus_1) + new_real_orders_count

    # Calculate the proportion of real orders for day t
    real_proportion_t = total_real_orders_t / total_orders

    # Generate orders for day t
    orders_t = generate_orders_for_day(real_proportion_t, 1 - real_proportion_t, existing_real_orders=real_orders_t_minus_1)

    # Perform allocation for day t-1 and day t
    allocation_t_minus_1 = allocate_orders(orders_t_minus_1, factory_capacities)
    allocation_t = allocate_orders(orders_t, factory_capacities)

    # Calculate WMAPE site and global before SA
    wmape_site_before_sa = calculate_wmape_site(allocation_t_minus_1, allocation_t, [])
    wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, [])

    # Apply Simulated Annealing with swapping to optimize allocation on day t
    optimized_allocation_t, optimized_wmape_site = simulated_annealing_with_swap(allocation_t_minus_1, allocation_t, factory_capacities)

    end_time = time.time()
    execution_time = end_time - start_time

    return execution_time, wmape_site_before_sa, optimized_wmape_site, wmape_global, factory_capacities

# Main execution code
order_quantities = list(range(1000, 10001, 1000))  # From 1000 to 10,000 in steps of 1000
execution_times = []
wmape_site_before_list = []
wmape_site_after_list = []
wmape_global_list = []
for quantity in order_quantities:
    print(f"Order quantity: {quantity}")
    exec_time, wmape_before, wmape_after, wmape_global, capacities = run_allocation_process(quantity)
    execution_times.append(exec_time)
    wmape_site_before_list.append(wmape_before)
    wmape_site_after_list.append(wmape_after)
    wmape_global_list.append(wmape_global)
    print(f"Execution time: {exec_time:.2f} seconds")
    print(f"WMAPE site before SA: {wmape_before:.3f}")
    print(f"WMAPE site after SA: {wmape_after:.3f}")
    print(f"WMAPE global: {wmape_global:.3f}")
    print("--------------------")

# Plotting execution time
plt.figure(figsize=(12, 8))
plt.plot(order_quantities, execution_times, marker='o')
plt.xlabel('Order quantity')
plt.ylabel('Execution time (seconds)')
plt.title('Order quantity vs Execution time')
plt.grid(True)
plt.show()

# Plotting WMAPE values
plt.figure(figsize=(12, 8))
plt.plot(order_quantities, wmape_site_before_list, marker='o', label='WMAPE site before SA')
plt.plot(order_quantities, wmape_site_after_list, marker='s', label='WMAPE site after SA')
plt.plot(order_quantities, wmape_global_list, marker='^', label='WMAPE global')
plt.xlabel('Order quantity')
plt.ylabel('WMAPE')
plt.title('Order quantity vs WMAPE')
plt.legend()
plt.grid(True)
plt.show()

# Export results to Excel
workbook = Workbook()
sheet = workbook.active
sheet.title = "Performance results"
sheet.append(["Order quantity", "Execution time (s)", "WMAPE site before SA", "WMAPE site after SA", "WMAPE global"])
for i, quantity in enumerate(order_quantities):
    sheet.append([
        quantity, 
        execution_times[i], 
        wmape_site_before_list[i], 
        wmape_site_after_list[i], 
        wmape_global_list[i]
    ])
workbook.save("performance_results.xlsx")