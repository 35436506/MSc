import random
import math
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

# Set seed for reproducibility
random.seed(42)

# Define the recipes with their eligibility
recipes_f1_only = list(range(1, 20))
recipes_f2_only = list(range(30, 66))
recipes_f1_f2 = list(range(20, 30))
all_recipes = recipes_f1_only + recipes_f1_f2 + recipes_f2_only

# Define total orders and factory capacities
total_orders = 50000
F1_cap = int(0.25 * total_orders)  # 25% of total orders
F2_cap = int(0.5 * total_orders)  # 50% of total orders
factory_capacities = {
    'F1': F1_cap,
    'F2': F2_cap,
    'F3': float('inf')  # F3 has unlimited capacity
}

def generate_order_recipes(eligible_recipes, num_recipes):
    return random.sample(eligible_recipes, random.randint(1, min(4, len(eligible_recipes))))

def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    orders = []
    f1_target = int(0.3 * total_orders)
    f2_target = int(0.6 * total_orders)
    f1_count = 0
    f2_count = 0
    total_real_orders = int(real_proportion * total_orders)
    
    # Handle existing real orders
    if existing_real_orders:
        for order in existing_real_orders:
            orders.append(order)
            if 'F1' in order['eligible_factories']:
                f1_count += 1
            if 'F2' in order['eligible_factories']:
                f2_count += 1

    # Generate Group 1 (F1 eligible) orders
    while f1_count < f1_target:
        recipe_ids = generate_order_recipes(all_recipes, 4)
        # Bias towards F1,F3 instead of F1,F2,F3
        eligible_factories = ['F1', 'F3'] if random.random() < 0.7 else ['F1', 'F2', 'F3']
        orders.append({
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        })
        f1_count += 1
        if 'F2' in eligible_factories:
            f2_count += 1

    # Generate Group 2 (F2 eligible) orders
    while f2_count < f2_target:
        recipe_ids = generate_order_recipes(all_recipes, 4)
        # Bias towards F2,F3 instead of F1,F2,F3
        if f1_count < f1_target and random.random() < 0.3:
            eligible_factories = ['F1', 'F2', 'F3']
            f1_count += 1
        else:
            eligible_factories = ['F2', 'F3']
        orders.append({
            'recipe_ids': recipe_ids,
            'eligible_factories': eligible_factories,
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        })
        f2_count += 1

    # Generate F3-only orders for the remainder
    while len(orders) < total_orders:
        recipe_ids = generate_order_recipes(all_recipes, 4)
        orders.append({
            'recipe_ids': recipe_ids,
            'eligible_factories': ['F3'],
            'is_real': len([o for o in orders if o['is_real']]) < total_real_orders
        })

    # Ensure exact number of real orders
    real_order_count = sum(1 for order in orders if order['is_real'])
    if real_order_count < total_real_orders:
        for order in random.sample([o for o in orders if not o['is_real']], total_real_orders - real_order_count):
            order['is_real'] = True
    elif real_order_count > total_real_orders:
        for order in random.sample([o for o in orders if o['is_real']], real_order_count - total_real_orders):
            order['is_real'] = False

    # Update order IDs
    for i, order in enumerate(orders, start=1):
        order['id'] = i

    return orders

def allocate_orders(orders, factory_capacities, day):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    # Allocate to F1 first, prioritizing F1,F3 and F1,F2,F3 orders
    f1_eligible = sorted([order for order in remaining_orders if 'F1' in order['eligible_factories']],
                         key=lambda x: len(x['eligible_factories']))  # Prioritize F1,F3 over F1,F2,F3
    allocation['F1'] = f1_eligible[:factory_capacities['F1']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F1']]

    # Allocate to F2, using remaining F1,F2,F3 orders and F2,F3 orders
    f2_eligible = [order for order in remaining_orders if 'F2' in order['eligible_factories']]
    allocation['F2'] = f2_eligible[:factory_capacities['F2']]
    remaining_orders = [order for order in remaining_orders if order not in allocation['F2']]

    # Allocate remaining to F3
    allocation['F3'] = remaining_orders

    # Verify and print allocation details
    print(f"\nAllocation details for Day {day}:")
    for factory, orders in allocation.items():
        print(f"{factory}: {len(orders)} orders allocated")
        print(f"  F1,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F1', 'F3'})}")
        print(f"  F2,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F2', 'F3'})}")
        print(f"  F1,F2,F3: {sum(1 for o in orders if set(o['eligible_factories']) == {'F1', 'F2', 'F3'})}")
        print(f"  F3 only: {sum(1 for o in orders if o['eligible_factories'] == ['F3'])}")

    return allocation

def plot_allocation(allocation, day):
    plt.figure(figsize=(12, 6))
    factories = ['F1', 'F2', 'F3']
    real_order_counts = [len([order for order in allocation[factory] if order['is_real']]) for factory in factories]
    simulated_order_counts = [len([order for order in allocation[factory] if not order['is_real']]) for factory in factories]

    x = np.arange(len(factories))
    width = 0.35

    plt.bar(x, real_order_counts, width, label='Real orders', color='#1f77b4')
    plt.bar(x, simulated_order_counts, width, bottom=real_order_counts, label='Simulated orders', color='#ff7f0e')

    max_capacity_line = None
    for i, factory in enumerate(factories):
        capacity = factory_capacities[factory]
        if factory == 'F3':
            capacity = real_order_counts[i] + simulated_order_counts[i] + 7000
        
        line_width = 0.6
        xmin = i - line_width/2
        xmax = i + line_width/2
        
        line = plt.hlines(y=capacity, xmin=xmin, xmax=xmax, 
                    colors='r', linestyles='-', linewidth=2)
        if max_capacity_line is None:
            max_capacity_line = line

    plt.xlabel('Factory')
    plt.ylabel('Order quantity')
    plt.title(f'Allocation for Day {day}')
    plt.xticks(x, factories)
    
    plt.legend([plt.Rectangle((0,0),1,1,fc="#1f77b4"), 
                plt.Rectangle((0,0),1,1,fc="#ff7f0e"),
                max_capacity_line],
               ['Real orders', 'Simulated orders', 'Max capacity'],
               bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.show()

def calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_items_t = 0
    
    # Get all unique recipe IDs from both days
    all_recipes = set()
    for allocation in [allocation_t_minus_1, allocation_t]:
        for factory in allocation:
            for order in allocation[factory]:
                all_recipes.update(order['recipe_ids'])
    
    # Initialize recipe counts for each factory on both days
    recipe_counts_t_minus_1 = {factory: {recipe: 0 for recipe in all_recipes} for factory in ['F1', 'F2', 'F3']}
    recipe_counts_t = {factory: {recipe: 0 for recipe in all_recipes} for factory in ['F1', 'F2', 'F3']}
    
    # Count the recipes for each factory on day t-1
    for factory, orders in allocation_t_minus_1.items():
        for order in orders:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[factory][recipe_id] += 1
    
    # Count the recipes for each factory on day t
    for factory, orders in allocation_t.items():
        for order in orders:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[factory][recipe_id] += 1
                total_items_t += 1
                
    # Create the table headers
    sheet_wmape.append(['Recipe', 'Factory', 'Day t-1', 'Day t', 'Absolute recipe-site error'])
    
    # Calculate the absolute difference for each recipe-site combination
    for recipe_id in all_recipes:
        for factory in ['F1', 'F2', 'F3']:
            t_minus_1_count = recipe_counts_t_minus_1[factory][recipe_id]
            t_count = recipe_counts_t[factory][recipe_id]
            abs_diff = abs(t_count - t_minus_1_count)
            total_abs_diff += abs_diff
            sheet_wmape.append([recipe_id, factory, t_minus_1_count, t_count, abs_diff])
    
    if total_items_t == 0:
        wmape_site = float('inf')
    else:
        wmape_site = total_abs_diff / total_items_t
        
    sheet_wmape.append(['', '', 'SUM', total_items_t, total_abs_diff])
    sheet_wmape.append(['WMAPE site', wmape_site])
    sheet_wmape.append([''])
    
    return wmape_site

def calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    
    # Count the recipes for each factory on day t-1
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    
    # Count the recipes for each factory on day t
    for factory in allocation_t:
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    
    # Get all unique recipe IDs from both days
    all_recipes = set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys())
    
    # Create the table headers
    sheet_wmape.append(['Recipe', 'Day t-1', 'Day t', 'Absolute recipe error'])
    
    # Calculate the absolute difference for each recipe
    for recipe_id in all_recipes:
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
        sheet_wmape.append([recipe_id, t_minus_1_count, t_count, abs_diff])
    
    # Add the total absolute difference and total items on day t to the table
    sheet_wmape.append(['', 'SUM', total_t_items, total_abs_diff])
    
    if total_t_items == 0:
        wmape_global = float('inf')
    else:
        wmape_global = total_abs_diff / total_t_items
    
    # Add the WMAPE global to the table
    sheet_wmape.append(['WMAPE global', wmape_global])
    
    return wmape_global

def simulated_annealing_with_swap(allocation_t_minus_1, allocation_t, factory_capacities, initial_temp=1000, cooling_rate=0.995, iterations=10000):
    current_allocation = {factory: orders[:] for factory, orders in allocation_t.items()}
    current_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, [])
    best_allocation = current_allocation.copy()
    best_wmape_site = current_wmape_site
    temp = initial_temp

    recipe_counts_t_minus_1 = get_recipe_counts(allocation_t_minus_1)

    for i in range(iterations):
        factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
        if current_allocation[factory1] and current_allocation[factory2]:
            idx1 = random.randint(0, len(current_allocation[factory1]) - 1)
            idx2 = random.randint(0, len(current_allocation[factory2]) - 1)
            order1 = current_allocation[factory1][idx1]
            order2 = current_allocation[factory2][idx2]
            
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                new_allocation = {f: orders[:] for f, orders in current_allocation.items()}
                new_allocation[factory1][idx1], new_allocation[factory2][idx2] = order2, order1
                
                new_wmape_site = calculate_wmape_site(allocation_t_minus_1, new_allocation, [])
                
                if new_wmape_site < current_wmape_site or random.random() < math.exp((current_wmape_site - new_wmape_site) / temp):
                    current_allocation = new_allocation
                    current_wmape_site = new_wmape_site
                    if current_wmape_site < best_wmape_site:
                        best_wmape_site = current_wmape_site
                        best_allocation = {f: orders[:] for f, orders in current_allocation.items()}
        
        # Local search
        if i % 100 == 0:
            current_allocation = local_search(allocation_t_minus_1, current_allocation, factory_capacities, recipe_counts_t_minus_1)
            current_wmape_site = calculate_wmape_site(allocation_t_minus_1, current_allocation, [])
            if current_wmape_site < best_wmape_site:
                best_wmape_site = current_wmape_site
                best_allocation = current_allocation.copy()

        temp *= cooling_rate

    return best_allocation, best_wmape_site

def get_recipe_counts(allocation):
    recipe_counts = {}
    for factory in allocation:
        for order in allocation[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts[recipe_id] = recipe_counts.get(recipe_id, 0) + 1
    return recipe_counts

def local_search(allocation_t_minus_1, current_allocation, factory_capacities, recipe_counts_t_minus_1):
    recipe_counts_current = get_recipe_counts(current_allocation)
    
    for l in range(100):  # Perform 100 local search iterations
        factory1, factory2 = random.sample(list(factory_capacities.keys()), 2)
        if current_allocation[factory1] and current_allocation[factory2]:
            order1 = max(current_allocation[factory1], key=lambda o: sum(abs(recipe_counts_current.get(r, 0) - recipe_counts_t_minus_1.get(r, 0)) for r in o['recipe_ids']))
            order2 = max(current_allocation[factory2], key=lambda o: sum(abs(recipe_counts_current.get(r, 0) - recipe_counts_t_minus_1.get(r, 0)) for r in o['recipe_ids']))
            
            if factory2 in order1['eligible_factories'] and factory1 in order2['eligible_factories']:
                idx1, idx2 = current_allocation[factory1].index(order1), current_allocation[factory2].index(order2)
                current_allocation[factory1][idx1], current_allocation[factory2][idx2] = order2, order1
                recipe_counts_current = get_recipe_counts(current_allocation)
    
    return current_allocation

# Main execution code
workbook = Workbook()

# Create sheets in the desired order
sheet_wmape = workbook.create_sheet("WMAPE", 0)
sheet_day_minus_15_orders = workbook.create_sheet("Day (-15) Orders", 1)
sheet_day_minus_15_allocation = workbook.create_sheet("Day (-15) Allocation", 2)
sheet_day_minus_14_orders = workbook.create_sheet("Day (-14) Orders", 3)
sheet_day_minus_14_allocation = workbook.create_sheet("Day (-14) Allocation (Initial)", 4)
sheet_day_minus_14_allocation_sa = workbook.create_sheet("Day (-14) Allocation (SA)", 5)

# Generate orders for day (-15)
orders_day_minus_15 = generate_orders_for_day(0.3, 0.7)  # 30% real orders

# Extract real orders from day (-15)
real_orders_day_minus_15 = [order for order in orders_day_minus_15 if order['is_real']]

# Generate orders for day (-14), including new real orders
orders_day_minus_14 = generate_orders_for_day(0.4, 0.6, existing_real_orders=real_orders_day_minus_15)  # 40% real orders

# Function to print order statistics
def print_order_statistics(orders, day):
    print(f"\nDay ({day}) Eligible Orders:")
    f1_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F3'})
    f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F2', 'F3'})
    f1_f2_f3_eligible = sum(1 for order in orders if set(order['eligible_factories']) == {'F1', 'F2', 'F3'})
    f3_only_eligible = sum(1 for order in orders if order['eligible_factories'] == ['F3'])
    print(f"Total orders with eligible factories F1, F3: {f1_f3_eligible}")
    print(f"Total orders with eligible factories F2, F3: {f2_f3_eligible}")
    print(f"Total orders with eligible factories F1, F2, F3: {f1_f2_f3_eligible}")
    print(f"Total orders with eligible factories F3: {f3_only_eligible}")
    print(f"Total F1 eligible: {f1_f3_eligible + f1_f2_f3_eligible}")
    print(f"Total F2 eligible: {f2_f3_eligible + f1_f2_f3_eligible}")
    print(f"Total orders: {len(orders)}")
    print(f"Total real orders: {sum(1 for order in orders if order['is_real'])}")

# Print statistics for both days
print_order_statistics(orders_day_minus_15, -15)
print_order_statistics(orders_day_minus_14, -14)

# Perform allocation for day (-15) and day (-14)
allocation_day_minus_15 = allocate_orders(orders_day_minus_15, factory_capacities, -15)
allocation_day_minus_14 = allocate_orders(orders_day_minus_14, factory_capacities, -14)

# Plot allocations
plot_allocation(allocation_day_minus_15, '(-15)')
plot_allocation(allocation_day_minus_14, '(-14) (Before SA)')

# Calculate WMAPE site and global before SA
wmape_site_before_sa = calculate_wmape_site(allocation_day_minus_15, allocation_day_minus_14, sheet_wmape)
wmape_global = calculate_wmape_global(allocation_day_minus_15, allocation_day_minus_14, sheet_wmape)

print(f"WMAPE site before SA: {wmape_site_before_sa:.3f}")
print(f"WMAPE global: {wmape_global:.3f}")

start_time = time.time()

# Apply Simulated Annealing with swapping to optimize allocation on day (-14)
optimized_allocation_day_minus_14, optimized_wmape_site = simulated_annealing_with_swap(allocation_day_minus_15, allocation_day_minus_14, factory_capacities)

end_time = time.time()
optimization_time = end_time - start_time

# Export WMAPE site after SA to Excel sheet
sheet_wmape.append([''])
sheet_wmape.append([f"WMAPE site after SA: {optimized_wmape_site:.3f}"])

# Plot solution allocation for day (-14) after SA
plot_allocation(optimized_allocation_day_minus_14, '(-14) (After SA)')

print(f"WMAPE site after SA: {optimized_wmape_site:.3f}")
print(f"Optimization time: {optimization_time:.2f} seconds")
print("-------------------------------")
print("                               ")

# Populate Day (-15) Orders sheet
sheet_day_minus_15_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_day_minus_15:
    sheet_day_minus_15_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day (-15) Allocation sheet
sheet_day_minus_15_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_day_minus_15.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_minus_15_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day (-14) Orders sheet
sheet_day_minus_14_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_day_minus_14:
    sheet_day_minus_14_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day (-14) Allocation (Initial) sheet
sheet_day_minus_14_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_day_minus_14.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_minus_14_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day (-14) Allocation (SA) sheet
sheet_day_minus_14_allocation_sa.append(["Factory", "Allocated Orders"])
for factory, orders in optimized_allocation_day_minus_14.items():
    order_ids = [order['id'] for order in orders]
    sheet_day_minus_14_allocation_sa.append([factory, ", ".join(map(str, order_ids))])

# Remove the default sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Save the workbook
workbook.save("allocation_results.xlsx")

# Print some statistics
print("Day (-15) Statistics:")
for factory, orders in allocation_day_minus_15.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nDay (-14) Statistics (Before SA):")
for factory, orders in allocation_day_minus_14.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nDay (-14) Statistics (After SA):")
for factory, orders in optimized_allocation_day_minus_14.items():
    print(f"{factory}: {len(orders)} orders, {sum(order['is_real'] for order in orders)} real")

print("\nWMAPE Site Improvement:")
print(f"Before SA: {wmape_site_before_sa:.3f}")
print(f"After SA: {optimized_wmape_site:.3f}")
print(f"Improvement: {(wmape_site_before_sa - optimized_wmape_site) / wmape_site_before_sa * 100:.2f}%")
