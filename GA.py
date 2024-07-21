import random
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

# Set seed for reproducibility
random.seed(42)

# Define the recipes as a list from 1 to 65
recipes = list(range(1, 66))

# Define the recipe eligibility for each factory
factory_eligibility = {
    'F1': random.sample(recipes, int(0.3 * len(recipes))),  # 30% of recipes
    'F2': random.sample(recipes, int(0.6 * len(recipes))),  # 60% of recipes
    'F3': recipes  # All recipes
}

# Define total orders and factory capacities
total_orders = 500
F1_cap = int(0.3 * total_orders)  # 30% of total orders
F2_cap = int(0.5 * total_orders)  # 50% of total orders
factory_capacities = {
    'F1': F1_cap,
    'F2': F2_cap,
    'F3': float('inf')  # F3 has unlimited capacity
}

def generate_orders(num_orders, is_real=True, factory=None, prioritize_f1_f2=False):
    orders = []
    while len(orders) < num_orders:
        if prioritize_f1_f2:
            if factory == 'F1':
                recipe_ids = random.sample(factory_eligibility['F1'], random.randint(1, 4))
            elif factory == 'F2':
                recipe_ids = random.sample(factory_eligibility['F2'], random.randint(1, 4))
            else:
                recipe_ids = random.sample(recipes, random.randint(1, 4))
        else:
            recipe_ids = random.sample(recipes, random.randint(1, 4))

        if factory is None or all(recipe_id in factory_eligibility[factory] for recipe_id in recipe_ids):
            order = {'id': len(orders) + 1, 'recipe_ids': recipe_ids, 'is_real': is_real}
            order['eligible_factories'] = get_eligible_factories(order)
            orders.append(order)
    return orders

def get_eligible_factories(order):
    return [factory for factory, eligible_recipes in factory_eligibility.items()
            if all(recipe_id in eligible_recipes for recipe_id in order['recipe_ids'])]

def generate_orders_for_day(real_proportion, simulated_proportion, existing_real_orders=None):
    num_real = int(real_proportion * total_orders)
    num_simulated = int(simulated_proportion * total_orders)
    
    if existing_real_orders:
        orders = existing_real_orders.copy()
        additional_real_orders = num_real - len(existing_real_orders)
        if additional_real_orders > 0:
            orders.extend(generate_orders(additional_real_orders))
    else:
        orders = generate_orders(num_real)
    
    orders.extend(generate_orders(num_simulated, is_real=False))
    
    for factory in ['F1', 'F2']:
        eligible_orders = [order for order in orders if factory in order['eligible_factories']]
        while len(eligible_orders) < factory_capacities[factory]:
            new_orders = generate_orders(factory_capacities[factory] - len(eligible_orders), 
                                         is_real=False, factory=factory, prioritize_f1_f2=True)
            orders.extend(new_orders)
            eligible_orders.extend(new_orders)
    
    for i, order in enumerate(orders):
        order['id'] = i + 1
    
    return orders

def allocate_orders(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    remaining_orders = orders.copy()

    for factory in ['F1', 'F2', 'F3']:
        capacity = factory_capacities[factory]
        eligible_orders = [order for order in remaining_orders if factory in order['eligible_factories']]
        eligible_orders.sort(key=lambda x: x['is_real'], reverse=True)  # Prioritize real orders
        
        if factory == 'F3':
            allocated_orders = eligible_orders  # Allocate all remaining orders to F3
        else:
            allocated_orders = eligible_orders[:int(capacity)]  # Convert capacity to int for slicing
        
        allocation[factory].extend(allocated_orders)
        remaining_orders = [order for order in remaining_orders if order not in allocated_orders]

        if factory == 'F2' and len(allocation[factory]) < capacity:
            # If F2 is not full, generate more orders
            additional_orders = generate_orders(int(capacity) - len(allocation[factory]), 
                                                is_real=False, factory='F2', prioritize_f1_f2=True)
            allocation[factory].extend(additional_orders)

    return allocation

def plot_allocation(allocation, day):
    plt.figure(figsize=(12, 6))
    factories = ['F1', 'F2', 'F3']
    real_order_counts = [len([order for order in allocation[factory] if order['is_real']]) for factory in factories]
    simulated_order_counts = [len([order for order in allocation[factory] if not order['is_real']]) for factory in factories]

    x = np.arange(len(factories))
    width = 0.35

    plt.bar(x, real_order_counts, width, label='Real orders', color='#1f77b4')
    plt.bar(x, simulated_order_counts, width, bottom=real_order_counts, label='Simulated orders', color='#ff7f0e')

    max_capacity_line = None
    for i, factory in enumerate(factories):
        capacity = factory_capacities[factory]
        if factory == 'F3':
            capacity = real_order_counts[i] + simulated_order_counts[i] + 1000
        
        line_width = 0.6
        xmin = i - line_width/2
        xmax = i + line_width/2
        
        line = plt.hlines(y=capacity, xmin=xmin, xmax=xmax, 
                    colors='r', linestyles='-', linewidth=2)
        if max_capacity_line is None:
            max_capacity_line = line

    plt.xlabel('Factory')
    plt.ylabel('Order quantity')
    plt.title(f'Allocation for Day {day}')
    plt.xticks(x, factories)
    
    plt.legend([plt.Rectangle((0,0),1,1,fc="#1f77b4"), 
                plt.Rectangle((0,0),1,1,fc="#ff7f0e"),
                max_capacity_line],
               ['Real orders', 'Simulated orders', 'Max capacity'],
               bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.tight_layout()
    plt.show()

def calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_items_t = 0
    
    for factory in ['F1', 'F2', 'F3']:
        factory_abs_diff = 0
        recipe_counts_t_minus_1 = {}
        recipe_counts_t = {}
        
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
        
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_items_t += 1
        
        sheet_wmape.append([f"Recipe counts for {factory} on day t-1: {recipe_counts_t_minus_1}"])
        sheet_wmape.append([f"Recipe counts for {factory} on day t: {recipe_counts_t}"])
        
        for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
            t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
            t_count = recipe_counts_t.get(recipe_id, 0)
            abs_diff = abs(t_count - t_minus_1_count)
            factory_abs_diff += abs_diff
            sheet_wmape.append([f"Recipe {recipe_id} in {factory}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
        
        total_abs_diff += factory_abs_diff
        
        sheet_wmape.append([f"Total absolute difference for {factory}: {factory_abs_diff}"])
        sheet_wmape.append([f"Total items for {factory} on day t: {sum(recipe_counts_t.values())}"])
        
    if total_items_t == 0:
        wmape_site = float('inf')
    else:
        wmape_site = total_abs_diff / total_items_t
    
    sheet_wmape.append([f"Total absolute difference across all factories: {total_abs_diff}"])
    sheet_wmape.append([f"Total items across all factories on day t: {total_items_t}"])
    sheet_wmape.append([f"WMAPE site: {wmape_site:.3f}"])
    
    return wmape_site

def calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    sheet_wmape.append([f"Recipe counts on day t-1: {recipe_counts_t_minus_1}"])
    for factory in allocation_t:
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    sheet_wmape.append([f"Recipe counts on day t: {recipe_counts_t}"])
    for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        abs_diff = abs(t_minus_1_count - t_count)
        total_abs_diff += abs_diff
        sheet_wmape.append([f"Recipe {recipe_id}: Day t-1 count = {t_minus_1_count}, Day t count = {t_count}, Absolute difference = {abs_diff}"])
    wmape_global = total_abs_diff / total_t_items
    sheet_wmape.append([f"Total absolute difference: {total_abs_diff}"])
    sheet_wmape.append([f"Total items on day t: {total_t_items}"])
    sheet_wmape.append([f"WMAPE global: {wmape_global:.4f}"])
    return wmape_global

def crossover(parent1, parent2):
    child = {}
    for factory in parent1:
        child[factory] = []
        for i in range(len(parent1[factory])):
            if random.random() < 0.5:
                child[factory].append(parent1[factory][i])
            else:
                child[factory].append(parent2[factory][i])
    return child

def mutate(individual, mutation_rate):
    for factory1 in individual:
        for i, order in enumerate(individual[factory1]):
            if random.random() < mutation_rate:
                factory2 = random.choice([f for f in individual if f != factory1])
                if individual[factory2]:
                    j = random.randint(0, len(individual[factory2]) - 1)
                    if factory2 in order['eligible_factories'] and factory1 in individual[factory2][j]['eligible_factories']:
                        individual[factory1][i], individual[factory2][j] = individual[factory2][j], individual[factory1][i]
    return individual

def genetic_algorithm(allocation_t_minus_1, allocation_t, factory_capacities, population_size=50, generations=100, mutation_rate=0.1):
    # Initialize population
    population = [allocation_t.copy() for _ in range(population_size)]
    
    best_individual = None
    best_fitness = float('inf')
    
    for generation in range(generations):
        # Evaluate fitness
        fitness_scores = [calculate_wmape_site(allocation_t_minus_1, individual, []) for individual in population]
        
        # Select best individual
        best_gen_individual = population[fitness_scores.index(min(fitness_scores))]
        best_gen_fitness = min(fitness_scores)
        
        if best_gen_fitness < best_fitness:
            best_individual = best_gen_individual
            best_fitness = best_gen_fitness
        
        # Select parents
        parents = random.choices(population, weights=[1/score for score in fitness_scores], k=population_size)
        
        # Create next generation
        next_generation = []
        for i in range(0, population_size, 2):
            parent1, parent2 = parents[i], parents[i+1]
            child1 = crossover(parent1, parent2)
            child2 = crossover(parent2, parent1)
            next_generation.extend([mutate(child1, mutation_rate), mutate(child2, mutation_rate)])
        
        population = next_generation
    
    return best_individual, best_fitness

# Main execution code
workbook = Workbook()

# Create sheets in the desired order
sheet_wmape = workbook.create_sheet("WMAPE", 0)
sheet_t_minus_1_orders = workbook.create_sheet("Day t-1 Orders", 1)
sheet_t_minus_1_allocation = workbook.create_sheet("Day t-1 Allocation", 2)
sheet_t_orders = workbook.create_sheet("Day t Orders", 3)
sheet_t_allocation = workbook.create_sheet("Day t Allocation (Initial)", 4)
sheet_t_allocation_ga = workbook.create_sheet("Day t Allocation (GA)", 5)

start_time = time.time()

# Generate orders for day t-1
orders_t_minus_1 = generate_orders_for_day(0.3, 0.7)

# Extract real orders from day t-1
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

# Calculate the number of new real orders for day t (10% of total orders)
new_real_orders_count = int(0.1 * total_orders)

# Calculate the total number of real orders for day t
total_real_orders_t = len(real_orders_t_minus_1) + new_real_orders_count

# Calculate the proportion of real orders for day t
real_proportion_t = total_real_orders_t / total_orders

# Generate orders for day t
orders_t = generate_orders_for_day(real_proportion_t, 1 - real_proportion_t, existing_real_orders=real_orders_t_minus_1)

# Perform allocation for day t-1 and day t
allocation_t_minus_1 = allocate_orders(orders_t_minus_1, factory_capacities)
allocation_t = allocate_orders(orders_t, factory_capacities)

# Plot allocations
plot_allocation(allocation_t_minus_1, 't-1')
plot_allocation(allocation_t, 't (Before GA)')

# Calculate WMAPE site and global before GA
wmape_site_before_ga = calculate_wmape_site(allocation_t_minus_1, allocation_t, sheet_wmape)
wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, sheet_wmape)

print(f"WMAPE site before GA: {wmape_site_before_ga:.3f}")
print(f"WMAPE global: {wmape_global:.3f}")

# Apply Genetic Algorithm to optimize allocation on day t
optimized_allocation_t, optimized_wmape_site = genetic_algorithm(allocation_t_minus_1, allocation_t, factory_capacities)

# Export WMAPE site after GA to Excel sheet
sheet_wmape.append([f"WMAPE site after GA: {optimized_wmape_site:.3f}"])

# Plot solution allocation for day t after GA
plot_allocation(optimized_allocation_t, 't (After GA)')

print(f"WMAPE site after GA: {optimized_wmape_site:.3f}")

end_time = time.time()
execution_time = end_time - start_time

print(f"Execution time: {execution_time:.2f} seconds")
print("-------------------------------")
print("                               ")

# Populate Day t-1 Orders sheet
sheet_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t_minus_1:
    sheet_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t-1 Allocation sheet
sheet_t_minus_1_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t_minus_1.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_minus_1_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Orders sheet
sheet_t_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t:
    sheet_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Populate Day t Allocation (Initial) sheet
sheet_t_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation.append([factory, ", ".join(map(str, order_ids))])

# Populate Day t Allocation (GA) sheet
sheet_t_allocation_ga.append(["Factory", "Allocated Orders"])
for factory, orders in optimized_allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation_ga.append([factory, ", ".join(map(str, order_ids))])

# Remove the default sheet if it exists
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Save the workbook
workbook.save("allocation_results.xlsx")

# Test scalability
def run_allocation_process(quantity):
    global total_orders, factory_capacities
    start_time = time.time()
    
    # Update total orders and factory capacities
    total_orders = quantity
    F1_cap = int(0.3 * total_orders)
    F2_cap = int(0.5 * total_orders)
    factory_capacities = {
        'F1': F1_cap,
        'F2': F2_cap,
        'F3': float('inf')
    }
    
    # Generate orders for day t-1
    orders_t_minus_1 = generate_orders_for_day(0.3, 0.7)

    # Extract real orders from day t-1
    real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]

    # Calculate the number of new real orders for day t (10% of total orders)
    new_real_orders_count = int(0.1 * total_orders)

    # Calculate the total number of real orders for day t
    total_real_orders_t = len(real_orders_t_minus_1) + new_real_orders_count

    # Calculate the proportion of real orders for day t
    real_proportion_t = total_real_orders_t / total_orders

    # Generate orders for day t
    orders_t = generate_orders_for_day(real_proportion_t, 1 - real_proportion_t, existing_real_orders=real_orders_t_minus_1)

    # Perform allocation for day t-1 and day t
    allocation_t_minus_1 = allocate_orders(orders_t_minus_1, factory_capacities)
    allocation_t = allocate_orders(orders_t, factory_capacities)

    # Calculate WMAPE site and global before GA
    wmape_site_before_ga = calculate_wmape_site(allocation_t_minus_1, allocation_t, [])
    wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t, [])

    # Apply Genetic Algorithm to optimize allocation on day t
    optimized_allocation_t, optimized_wmape_site = genetic_algorithm(allocation_t_minus_1, allocation_t, factory_capacities)

    end_time = time.time()
    execution_time = end_time - start_time

    return execution_time, wmape_site_before_ga, optimized_wmape_site, wmape_global, factory_capacities

# Main execution code for scalability test
order_quantities = list(range(1000, 10001, 1000))  # From 1000 to 10,000 in steps of 1000
execution_times = []
wmape_site_before_list = []
wmape_site_after_list = []
wmape_global_list = []
for quantity in order_quantities:
    print(f"Order quantity: {quantity}")
    exec_time, wmape_before, wmape_after, wmape_global, capacities = run_allocation_process(quantity)
    execution_times.append(exec_time)
    wmape_site_before_list.append(wmape_before)
    wmape_site_after_list.append(wmape_after)
    wmape_global_list.append(wmape_global)
    print(f"Execution time: {exec_time:.2f} seconds")
    print(f"WMAPE site before GA: {wmape_before:.3f}")
    print(f"WMAPE site after GA: {wmape_after:.3f}")
    print(f"WMAPE global: {wmape_global:.3f}")
    print("--------------------")

# Plotting execution time
plt.figure(figsize=(12, 8))
plt.plot(order_quantities, execution_times, marker='o')
plt.xlabel('Order quantity')
plt.ylabel('Execution time (seconds)')
plt.title('Order quantity vs Execution time')
plt.grid(True)
plt.show()

# Plotting WMAPE values
plt.figure(figsize=(12, 8))
plt.plot(order_quantities, wmape_site_before_list, marker='o', label='WMAPE site before GA')
plt.plot(order_quantities, wmape_site_after_list, marker='s', label='WMAPE site after GA')
plt.plot(order_quantities, wmape_global_list, marker='^', label='WMAPE global')
plt.xlabel('Order quantity')
plt.ylabel('WMAPE')
plt.title('Order quantity vs WMAPE')
plt.legend()
plt.grid(True)
plt.show()

# Export results to Excel
workbook = Workbook()
sheet = workbook.active
sheet.title = "Performance results"
sheet.append(["Order quantity", "Execution time (s)", "WMAPE site before GA", "WMAPE site after GA", "WMAPE global"])
for i, quantity in enumerate(order_quantities):
    sheet.append([
        quantity, 
        execution_times[i], 
        wmape_site_before_list[i], 
        wmape_site_after_list[i], 
        wmape_global_list[i]
    ])
workbook.save("performance_results.xlsx")