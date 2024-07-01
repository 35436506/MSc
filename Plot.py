import random
from openpyxl import Workbook
import matplotlib.pyplot as plt

# Define the recipes as a list from 1 to 10
recipes = list(range(1, 11))

# Define the recipe eligibility for each factory
factory_eligibility = {
    'F1': random.sample(recipes, int(0.5 * len(recipes))),  # 50% of total recipes are eligible for F1
    'F2': random.sample(recipes, int(0.8 * len(recipes))),  # 70% of total recipes are eligible for F2
    'F3': recipes  # 100% recipes are eligible for F3
}

# Generate random orders with 1-4 recipes
def generate_orders(num_orders, is_real=True):
    orders = []
    for n in range(num_orders):
        recipe_ids = random.sample(recipes, random.randint(1, 4))
        orders.append({'id': len(orders) + 1, 'recipe_ids': recipe_ids, 'is_real': is_real})
    return orders

# Get the eligible factories for each order based on the recipes
def get_eligible_factories(order):
    eligible_factories = []
    for factory, eligible_recipes in factory_eligibility.items():
        if all(recipe_id in eligible_recipes for recipe_id in order['recipe_ids']):
            eligible_factories.append(factory)
    return eligible_factories

# Generate sample orders for day t-1 and day t
num_total_orders = 60
num_real_orders_t_minus_1 = num_total_orders // 2
num_simulated_orders_t_minus_1 = num_total_orders - num_real_orders_t_minus_1
num_new_real_orders_t = num_total_orders // 6
num_simulated_orders_t = num_total_orders // 3

orders_t_minus_1 = generate_orders(num_real_orders_t_minus_1) + generate_orders(num_simulated_orders_t_minus_1, is_real=False)
orders_t = orders_t_minus_1[:num_real_orders_t_minus_1] + generate_orders(num_new_real_orders_t) + generate_orders(num_simulated_orders_t, is_real=False)

# Re-number the orders for day t-1 and day t
for i, order in enumerate(orders_t_minus_1):
    order['id'] = i + 1

for i, order in enumerate(orders_t):
    order['id'] = i + 1

# Add eligible factories to each order
for order in orders_t_minus_1 + orders_t:
    order['eligible_factories'] = get_eligible_factories(order)

# Define factory capacities
factory_capacities = {
    'F1': int((num_total_orders) / 3),  # Serve a third of total orders
    'F2': int((num_total_orders) / 2),  # Serve half of total orders
    'F3': float('inf')  # Catch-all factory with infinite capacity
}

# Print the maximum capacity of each factory
print("Maximum Capacity of Each Factory:")
for factory, capacity in factory_capacities.items():
    print(f"{factory}: {capacity}")

# Initial allocation (real orders are prioritized from F1 to F2 to F3)
def initial_allocation(orders, factory_capacities):
    allocation = {factory: [] for factory in factory_capacities}
    
    # Allocate real orders first
    real_orders = [order for order in orders if order['is_real']]
    for factory in ['F1', 'F2', 'F3']:
        for order in real_orders:
            if factory in order['eligible_factories'] and len(allocation[factory]) < factory_capacities[factory]:
                if all(recipe_id in factory_eligibility[factory] for recipe_id in order['recipe_ids']):
                    allocation[factory].append(order)
                    real_orders.remove(order)
    
    # Allocate remaining orders (simulated orders and any unallocated real orders)
    remaining_orders = [order for order in orders if order not in allocation['F1'] and order not in allocation['F2'] and order not in allocation['F3']]
    for order in remaining_orders:
        for factory in ['F1', 'F2', 'F3']:
            if len(allocation[factory]) < factory_capacities[factory]:
                allocation[factory].append(order)
                break
    
    return allocation

# Perform initial allocation for day t-1 and day t
allocation_t_minus_1 = initial_allocation(orders_t_minus_1, factory_capacities)
allocation_t = initial_allocation(orders_t, factory_capacities)

# Plot solution allocation for day t-1
plt.figure(figsize=(8, 4))
real_orders_t_minus_1 = [order for order in orders_t_minus_1 if order['is_real']]
simulated_orders_t_minus_1 = [order for order in orders_t_minus_1 if not order['is_real']]
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t_minus_1['F1'] if order['is_real']]),
                             len([order for order in allocation_t_minus_1['F2'] if order['is_real']]),
                             len([order for order in allocation_t_minus_1['F3'] if order['is_real']])],
        color='blue', label='Real Orders')
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t_minus_1['F1'] if not order['is_real']]),
                             len([order for order in allocation_t_minus_1['F2'] if not order['is_real']]),
                             len([order for order in allocation_t_minus_1['F3'] if not order['is_real']])],
        color='yellow', label='Simulated Orders', bottom=[len([order for order in allocation_t_minus_1[factory] if order['is_real']]) for factory in ['F1', 'F2', 'F3']])
plt.xlabel('Factory')
plt.ylabel('Volume')
plt.title('Allocation for Day t-1')
plt.legend()
plt.show()

# Plot solution allocation for day t
plt.figure(figsize=(8, 4))
real_orders_t = [order for order in orders_t if order['is_real']]
simulated_orders_t = [order for order in orders_t if not order['is_real']]
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t['F1'] if order['is_real']]),
                             len([order for order in allocation_t['F2'] if order['is_real']]),
                             len([order for order in allocation_t['F3'] if order['is_real']])],
        color='blue', label='Real Orders')
plt.bar(['F1', 'F2', 'F3'], [len([order for order in allocation_t['F1'] if not order['is_real']]),
                             len([order for order in allocation_t['F2'] if not order['is_real']]),
                             len([order for order in allocation_t['F3'] if not order['is_real']])],
        color='yellow', label='Simulated Orders', bottom=[len([order for order in allocation_t[factory] if order['is_real']]) for factory in ['F1', 'F2', 'F3']])
plt.xlabel('Factory')
plt.ylabel('Volume')
plt.title('Allocation for Day t')
plt.legend()
plt.show()

# Calculate WMAPE site
def calculate_wmape_site(allocation_t_minus_1, allocation_t):
    total_abs_diff = 0
    total_t_items = 0
    for factory in ['F1', 'F2']:
        recipe_counts_t_minus_1 = {}
        recipe_counts_t = {}
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
        for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
            t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
            t_count = recipe_counts_t.get(recipe_id, 0)
            total_abs_diff += abs(t_minus_1_count - t_count)
    wmape_site = total_abs_diff / total_t_items
    return wmape_site

# Calculate WMAPE global
def calculate_wmape_global(allocation_t_minus_1, allocation_t):
    total_abs_diff = 0
    total_t_items = 0
    recipe_counts_t_minus_1 = {}
    recipe_counts_t = {}
    for factory in allocation_t_minus_1:
        for order in allocation_t_minus_1[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t_minus_1[recipe_id] = recipe_counts_t_minus_1.get(recipe_id, 0) + 1
    for factory in allocation_t:
        for order in allocation_t[factory]:
            for recipe_id in order['recipe_ids']:
                recipe_counts_t[recipe_id] = recipe_counts_t.get(recipe_id, 0) + 1
                total_t_items += 1
    for recipe_id in set(recipe_counts_t_minus_1.keys()) | set(recipe_counts_t.keys()):
        t_minus_1_count = recipe_counts_t_minus_1.get(recipe_id, 0)
        t_count = recipe_counts_t.get(recipe_id, 0)
        total_abs_diff += abs(t_minus_1_count - t_count)
    wmape_global = total_abs_diff / total_t_items
    return wmape_global

# Calculate WMAPE site and global
wmape_site = calculate_wmape_site(allocation_t_minus_1, allocation_t)
wmape_global = calculate_wmape_global(allocation_t_minus_1, allocation_t)

print(f"WMAPE Site: {wmape_site:.2f}")
print(f"WMAPE Global: {wmape_global:.2f}")

# Export data and allocation decisions to Excel
workbook = Workbook()

# Export orders for day t-1
sheet_t_minus_1_orders = workbook.create_sheet("Day t-1 Orders")
sheet_t_minus_1_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t_minus_1:
    sheet_t_minus_1_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Export allocation for day t-1
sheet_t_minus_1_allocation = workbook.create_sheet("Day t-1 Allocation")
sheet_t_minus_1_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t_minus_1.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_minus_1_allocation.append([factory, ", ".join(map(str, order_ids))])

# Export orders for day t
sheet_t_orders = workbook.create_sheet("Day t Orders")
sheet_t_orders.append(["Order ID", "Recipe IDs", "Is Real", "Eligible Factories"])
for order in orders_t:
    sheet_t_orders.append([order['id'], ", ".join(map(str, order['recipe_ids'])), "Yes" if order['is_real'] else "No", ", ".join(order['eligible_factories'])])

# Export allocation for day t
sheet_t_allocation = workbook.create_sheet("Day t Allocation")
sheet_t_allocation.append(["Factory", "Allocated Orders"])
for factory, orders in allocation_t.items():
    order_ids = [order['id'] for order in orders]
    sheet_t_allocation.append([factory, ", ".join(map(str, order_ids))])

workbook.save("allocation_results.xlsx")